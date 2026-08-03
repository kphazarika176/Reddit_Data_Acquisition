import os
from datetime import datetime
from typing import Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from src.database import DatabaseManager
from src.logger import get_logger

logger = get_logger(__name__)


def flatten_dict(d: Dict[str, Any], parent_key: str = "", sep: str = "_") -> Dict[str, Any]:
    """Helper function to recursively flatten a dictionary."""
    items = []

    for k, v in d.items():
        new_key = f"{parent_key}{sep}{k}" if parent_key else k

        if isinstance(v, dict):
            items.extend(flatten_dict(v, new_key, sep=sep).items())
        else:
            items.append((new_key, v))

    return dict(items)


def format_value(val: Any) -> Any:
    """Formats values for Excel cells."""
    if val is None:
        return ""

    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d %H:%M:%S")

    return val


def export_to_excel(output_filename: str = "reddit_data_export.xlsx") -> str:
    """
    Reads data from SQLite tables (posts, comments, qa_pairs)
    and exports them to Excel.
    """

    logger.info("Initializing Excel export...")

    db = DatabaseManager()

    wb = Workbook()

    default_sheet = wb.active
    if default_sheet is not None:
        wb.remove(default_sheet)

    export_configs = [
        {
            "data": db.get_all_posts(),
            "sheet_name": "Reddit Posts",
            "columns": [
                "post_id",
                "subreddit",
                "title",
                "body",
                "author",
                "score",
                "num_comments",
                "sentiment_label",
                "sentiment_score",
                "keywords",
                "url",
                "created_utc",
            ],
        },
        {
            "data": db.get_latest_comments(1000000),
            "sheet_name": "Reddit Comments",
            "columns": [
                "comment_id",
                "post_id",
                "parent_id",
                "author",
                "body",
                "score",
                "depth",
                "sentiment_label",
                "sentiment_score",
                "keywords",
                "created_utc",
            ],
        },
        {
            "data": db.get_latest_qa_pairs(1000000),
            "sheet_name": "Q&A Pairs",
            "columns": [
                "question_comment_id",
                "answer_comment_id",
                "post_id",
                "question",
                "answer",
                "score_signal",
                "match_type",
                "verification_status",
                "confidence_score",
            ],
        },
    ]

    header_font = Font(name="Calibri", size=11, bold=True, color="000000")
    header_fill = PatternFill(
        start_color="E6EDF2",
        end_color="E6EDF2",
        fill_type="solid",
    )
    cell_font = Font(name="Calibri", size=11)

    counts = {}

    for config in export_configs:

        documents = config["data"]
        sheet_name = config["sheet_name"]
        fixed_cols = config["columns"]

        logger.info(f"Exporting worksheet '{sheet_name}'...")

        ws = wb.create_sheet(title=sheet_name)

        ws.freeze_panes = "A2"

        counts[sheet_name] = len(documents)

        header_names = [c.replace("_", " ").title() for c in fixed_cols]
        ws.append(header_names)

        for col_idx in range(1, len(header_names) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="left", vertical="center")

        for doc in documents:

            flat_doc = flatten_dict(doc)

            row = [
                format_value(flat_doc.get(col_key, ""))
                for col_key in fixed_cols
            ]

            ws.append(row)

            row_idx = ws.max_row

            for col_idx in range(1, len(fixed_cols) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = cell_font
                cell.alignment = Alignment(horizontal="left", vertical="center")

        max_col_letter = get_column_letter(len(fixed_cols))
        ws.auto_filter.ref = f"A1:{max_col_letter}{len(documents)+1}"

        for column in ws.columns:

            max_len = 0

            for cell in column:
                value = str(cell.value or "")
                max_len = max(max_len, len(value))

            width = min(max(max_len + 3, 12), 50)

            ws.column_dimensions[get_column_letter(column[0].column)].width = width

    project_root = os.path.abspath(
        os.path.join(os.path.dirname(__file__), "..")
    )

    filepath = os.path.join(project_root, output_filename)

    try:
        wb.save(filepath)
    except PermissionError:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base, ext = os.path.splitext(output_filename)
        fallback_filename = f"{base}_{timestamp}{ext}"
        filepath = os.path.join(project_root, fallback_filename)
        wb.save(filepath)
        print(f"\n[NOTE] File '{output_filename}' is currently open in Excel.")
        print(f"[NOTE] Saved export to fallback file: '{fallback_filename}'")

    print("\n====== SQLite to Excel Export Summary ======")

    for sheet, count in counts.items():
        print(f"  {sheet:<16}: {count} records exported")

    print(f"Workbook successfully saved to: {filepath}\n")

    logger.info(f"Excel export completed successfully. Saved to: {filepath}")

    db.close()

    return filepath